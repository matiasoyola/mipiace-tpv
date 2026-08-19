# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import unicodedata, re
from collections import defaultdict
from data import db, carta

def asc(s): return unicodedata.normalize('NFKD',s).encode('ascii','ignore').decode().lower().strip()
def norm(s):
    x=asc(s); x=re.sub(r'desayuno\s*\d+','',x); x=re.sub(r'[^a-z0-9 ]',' ',x)
    return re.sub(r'\s+',' ',x).strip()

# duplicados globales (para flags en categorias raw)
groups=defaultdict(list)
for i,(n,p) in enumerate(db): groups[norm(n)].append(i)

# ---- Canonicas (menu correcto validado con el usuario) ----
DES_CARTA={1:2.90,2:2.90,3:2.80,4:3.00,5:3.90,6:3.10,7:3.10,8:2.90,9:3.20,10:4.20}
DES_NAME={
 1:"1 - Cafe + Tostada mantequilla y mermelada",2:"2 - Cafe + Tostada tomate y aceite",
 3:"3 - Cafe + Croissant",4:"4 - Cafe + Croissant a la plancha",
 5:"5 - Cafe + Croissant o sandwich mixto",6:"6 - Cola Cao + Tostada mantequilla y mermelada",
 7:"7 - Cola Cao + Tostada tomate y aceite",8:"8 - Cola Cao + Croissant",
 9:"9 - Cola Cao + Croissant a la plancha",10:"10 - Cola Cao + Croissant o sandwich mixto"}
# precio BD actual de la serie "Desayuno N"
DES_BD={}
for n,p in db:
    m=re.match(r'Desayuno (\d+)',n)
    if m: DES_BD[int(m.group(1))]=p

canon_desayunos=[(DES_NAME[k], DES_BD.get(k), DES_CARTA[k],
                  ("Estaba a 0,00 - CORREGIR" if DES_BD.get(k)==0 else "Actualizar precio")) for k in range(1,11)]

canon_suplementos=[
 ("Tarrina de mermelada", None, 0.30, "Falta en BD - CREAR a 0,30"),
 ("Tarrina de mantequilla", None, 0.30, "Falta en BD - CREAR a 0,30"),
 ("Tarrina de tomate", None, 0.30, "Falta en BD - CREAR a 0,30"),
 ("Tarrina de aceite", None, 0.30, "Falta en BD - CREAR a 0,30"),
]

canon_cafe=[
 ("Cafe solo", 1.40, 1.50, "Actualizar a 1,50"),
 ("Cafe con leche", None, 1.50, "BD: 1,40 y 1,50 (2 duplicados) -> dejar 1 a 1,50"),
 ("Cafe cortado", None, 1.50, "BD: 1,40 y 1,50 (2 duplicados) -> dejar 1 a 1,50"),
 ("Cafe americano", None, 1.50, "BD: 1,40 y 1,50 (2 duplicados) -> dejar 1 a 1,50"),
 ("Cafe bombon", 2.00, 2.00, "BD: 2 duplicados a 2,00 -> dejar 1"),
 ("Capuccino", None, 1.90, "BD: Capuccino 2,00 / Capuchino 1,90 (2 dup) -> dejar 1 a 1,90"),
 ("Carajillo", 2.00, 2.00, "BD: 2 duplicados a 2,00 -> dejar 1"),
 ("Vaso de leche", 1.30, 1.50, "Actualizar a 1,50"),
 ("Cola Cao / Nesquik", 1.60, 1.70, "Actualizar a 1,70"),
 ("Chocolate", None, 2.40, "Falta en BD - CREAR a 2,40"),
]

canon_infusiones=[
 ("Infusion (manzanilla, poleo menta, tila, te verde, te rojo)", 1.40, 1.50,
  "Agrupa 5 sabores. Actualizar a 1,50. Sobran en BD: Te Verde / Te Rojo / Te de Tila / Infusion Clasica"),
]
canon_tefrio=[
 ("Te frio (limon, frutos bosque, roibos tropical, melocoton, pina colada, menta, melon, mojito)", None, 2.40,
  "BD: 'Te Frio' 2,50 y 'Te frio' 2,20 (2 duplicados) -> dejar 1 a 2,40"),
]

canon_bolleria=[
 ("Croissant", 1.50, 1.70, "BD 'Croissant Clasico' 1,50 -> 1,70"),
 ("Croissant nutella", 2.00, 2.50, "Actualizar a 2,50"),
 ("Tortitas con sirope o nutella", None, 2.70, "Normalizar en BD (hoy Tortitas Dulces/Variadas)"),
 ("Tortitas con nata", 2.50, 2.70, "Actualizar a 2,70"),
 ("Tortitas con sirope y nata", None, 3.00, "Crear / normalizar"),
 ("Tortitas con sirope y helado", None, 3.00, "Crear / normalizar"),
 ("Tortitas con sirope, nata y helado", None, 3.50, "Crear / normalizar"),
 ("Goffre con sirope o nutella", None, 3.70, "Normalizar (hoy Gofre/Goffre Especial/Irresistible)"),
 ("Goffre con nata", None, 3.70, "Crear / normalizar"),
 ("Goffre con sirope y nata", None, 4.00, "Crear / normalizar"),
 ("Goffre con sirope y helado", None, 4.00, "Crear / normalizar"),
 ("Goffre con sirope, nata y helado", None, 4.50, "Crear / normalizar"),
 ("Donut Glace", 1.32, 1.50, "BD 'Donut' 1,32 -> 1,50"),
 ("Donut Chocolate", None, 1.60, "Falta en BD - CREAR a 1,60"),
 ("Napolitana", 2.00, 1.90, "Bajar a 1,90"),
 ("Porcion bizcocho", None, 2.90, "Falta en BD - CREAR a 2,90"),
 ("Porcion Tarta", None, 3.60, "Falta en BD - CREAR a 3,60"),
 ("Palmera chocolate pequena", None, 1.40, "BD: 2 palmeras (2,00 y 2,20) -> normalizar; pequena 1,40"),
 ("Palmera grande", None, 2.80, "Falta en BD - CREAR a 2,80"),
]

canon_tostada=[
 ("Mermelada y mantequilla", 1.40, 1.90, "Actualizar a 1,90"),
 ("Tomate y aceite", 1.40, 1.90, "BD 'tomate' -> renombrar; a 1,90"),
 ("Serrano con tomate", 2.60, 3.10, "Actualizar a 3,10"),
 ("Iberico con tomate", 3.50, 4.20, "Actualizar a 4,20"),
 ("York y queso", 2.50, 3.00, "Actualizar a 3,00"),
 ("Atun, tomate y mahonesa", None, 3.10, "Falta en BD - CREAR a 3,10"),
 ("Guacamole", 3.00, 3.40, "Actualizar a 3,40"),
 ("Guacamole con salmon", 3.50, 4.60, "Actualizar a 4,60"),
 ("Guacamole con jamon iberico", 4.00, 4.60, "Actualizar a 4,60"),
 ("Salmon y alioli", 3.50, 4.20, "Actualizar a 4,20"),
]

canon_crsand=[
 ("Croissant plancha", None, 1.90, "BD: 2 duplicados (1,71 y 1,50) -> dejar 1 a 1,90"),
 ("Croissant mixto", None, 3.00, "BD: 2 duplicados (2,70 y 2,40) -> dejar 1 a 3,00"),
 ("Sandwich mixto", None, 3.00, "BD: 2 duplicados (2,70 y 2,40) -> dejar 1 a 3,00"),
 ("Sandwich vegetal", None, 3.10, "BD: 2 duplicados (2,70 y 2,60) -> dejar 1 a 3,10"),
 ("Sandwich vegetal con atun", 2.60, 3.40, "Actualizar a 3,40; revisar 'Sandwich Atun' 2,90 (sobra?)"),
]

canon_bebidas=[
 ("Refrescos", None, 2.50, "BD: 'Refresco' 2,40 y 'Refrescos' 2,20 -> unificar a 2,50"),
 ("Zumos", None, 2.50, "BD: 'Zumo' 2,20 y 'Zumo Natural' 2,30 -> unificar a 2,50"),
 ("Agua mineral", 1.00, 1.00, "Se queda a 1,00 (keeper 'Agua Mineral')"),
 ("Tinto de verano", None, 2.50, "BD: 2 duplicados (1,50 y 2,20) -> dejar 1 a 2,50"),
 ("Botellin", 1.40, 1.50, "Actualizar a 1,50"),
 ("Tercio", 2.00, 2.50, "Actualizar a 2,50; revisar 'Tercio 1906' 2,50"),
 ("Cerveza 1906", None, 2.90, "BD: 'Cerveza Artesanal' 2,70 / 'Tercio 1906' 2,50 -> normalizar a 2,90"),
 ("Alhambra", None, 2.90, "Falta en BD - CREAR a 2,90"),
 ("Botellin sin gluten", None, 1.70, "Falta en BD - CREAR a 1,70"),
 ("Batido chocolate", None, 2.50, "BD: 'Batido' 2,50 y 2,30 -> unificar a 2,50 ('Batidos naturales' es otro)"),
]
canon_licores=[
 ("Martini", 2.50, 3.00, "Actualizar a 3,00"),
 ("Baileys", None, 3.50, "BD: 2 duplicados (3,00 y 3,19) -> dejar 1 a 3,50"),
 ("Copa licor", None, 2.90, "BD: 2 duplicados (2,70 y 2,50) -> dejar 1 a 2,90"),
 ("Chupitos", None, 1.60, "BD: 'Chupito' 1,29 y 'Chipito licor' 1,20 -> unificar a 1,60"),
 ("Combinados", 5.00, 5.00, "OK, sin cambio"),
 ("Copa Whisky", None, 3.60, "BD: 2 duplicados (3,61 y 3,50) -> dejar 1 a 3,60"),
 ("Copa Soberano", None, 2.50, "BD: 2 duplicados (2,40 y 2,50) -> dejar 1 a 2,50"),
 ("Copa Castellana", None, 2.50, "Falta en BD - CREAR a 2,50"),
]

# ---- Ruteo de filas RAW (categorias aun sin canonizar y cubos de limpieza) ----
CAFE_KW=['cafe','solo','cortado','americano','bombon','capuc','carajillo','vaso de leche','colacao','nesquik']
def route(name):
    l=asc(name)
    if re.match(r'desayuno \d+',l): return None            # sustituido por canonico
    if l.startswith('tarrina de'): return None             # no existe en BD
    if l.startswith('cafe +') or l.startswith('colacao +'): return "DESAYUNOS ANTIGUOS (a eliminar)"
    if l.startswith('suplemento') or 'complementa tu desayuno' in l: return "ZUMO como suplemento (a eliminar)"
    if l.startswith('tostada o mollete'): return None
    if 'sandwich' in l: return "CROISSANT / SANDWICH actual en BD (limpiar)"
    if l.startswith('croissant'): return "CROISSANT / SANDWICH actual en BD (limpiar)" if ('plancha' in l or 'mixto' in l) else "BOLLERIA actual en BD (limpiar)"
    if l=='cucurucho' or l.startswith('tarrina'): return "HELADOS"
    if any(k in l for k in ['infusion','te frio','te rojo','te verde','te de tila']): return "INFUSIONES / TE actual en BD (limpiar)"
    if any(k in l for k in ['baileys','chupito','chipito','combinado','copa','martini']): return "LICORES actual en BD (limpiar)"
    if any(k in l for k in ['agua','botellin','cerveza','refresco','tercio','tinto','zumo','batido','alhambra']): return "BEBIDAS actual en BD (limpiar)"
    if any(k in l for k in ['donut','napolitana','palmera','goffre','gofre','tortitas']): return "BOLLERIA actual en BD (limpiar)"
    if 'tpv' in l or 'linea libre' in l or l=='distinto': return "OTROS / TPV"
    if any(k in l for k in CAFE_KW): return "CAFE ACTUAL EN BD (limpiar duplicados)"
    return "OTROS / TPV"

def raw_note(n,p):
    if n.startswith("TPV"): return ("Linea libre (0 EUR es correcto)","ok")
    if p==0.0: return ("Precio 0,00 - ROTO","bad")
    if len(groups[norm(n)])>1: return ("Duplicado en BD","warn")
    return ("","")

# Construir estructura display: cat -> list de (name, bd, carta, note, kind)
ORDER=["DESAYUNOS","SUPLEMENTOS","CAFE","INFUSIONES","TE FRIO","BOLLERIA","TOSTADA / MOLLETE",
       "CROISSANT / SANDWICH","BEBIDAS","LICORES","HELADOS",
       "DESAYUNOS ANTIGUOS (a eliminar)","CAFE ACTUAL EN BD (limpiar duplicados)",
       "BOLLERIA actual en BD (limpiar)","CROISSANT / SANDWICH actual en BD (limpiar)","BEBIDAS actual en BD (limpiar)","LICORES actual en BD (limpiar)","INFUSIONES / TE actual en BD (limpiar)","ZUMO como suplemento (a eliminar)","OTROS / TPV"]
data=defaultdict(list)

def kind_of(note):
    u=note.upper()
    if "ROTO" in u: return "bad"
    if "ELIMINAR" in u: return "bad"
    if "->" in u or any(w in u for w in ["CREAR","DUPLIC","ACTUALIZAR","CORREGIR","DEJAR","NORMALIZAR","BAJAR","SUBIR"]): return "warn"
    if "0 EUR ES CORRECTO" in u: return "ok"
    return ""

for name,bd,ct,note in canon_desayunos: data["DESAYUNOS"].append((name,bd,ct,note,kind_of(note)))
for name,bd,ct,note in canon_suplementos: data["SUPLEMENTOS"].append((name,bd,ct,note,kind_of(note)))
for name,bd,ct,note in canon_cafe: data["CAFE"].append((name,bd,ct,note,kind_of(note)))
for name,bd,ct,note in canon_infusiones: data["INFUSIONES"].append((name,bd,ct,note,kind_of(note)))
for name,bd,ct,note in canon_tefrio: data["TE FRIO"].append((name,bd,ct,note,kind_of(note)))
for name,bd,ct,note in canon_bolleria: data["BOLLERIA"].append((name,bd,ct,note,kind_of(note)))
for name,bd,ct,note in canon_tostada: data["TOSTADA / MOLLETE"].append((name,bd,ct,note,kind_of(note)))
for name,bd,ct,note in canon_crsand: data["CROISSANT / SANDWICH"].append((name,bd,ct,note,kind_of(note)))
for name,bd,ct,note in canon_bebidas: data["BEBIDAS"].append((name,bd,ct,note,kind_of(note)))
for name,bd,ct,note in canon_licores: data["LICORES"].append((name,bd,ct,note,kind_of(note)))

for n,p in db:
    c=route(n)
    if c is None: continue
    note,kind=raw_note(n,p)
    if c.endswith("(a eliminar)") and kind not in("bad",): 
        note = note or "Revisar / eliminar"; kind="warn"
    data[c].append((n,p,None,note,kind))

# ---- Excel ----
FONT="Arial"
hdr_fill=PatternFill('solid',fgColor='1F4E78'); hdr_font=Font(name=FONT,bold=True,color='FFFFFF',size=11)
catf=PatternFill('solid',fgColor='2E75B6'); delf=PatternFill('solid',fgColor='C00000')
catfont=Font(name=FONT,bold=True,color='FFFFFF',size=11)
bad_fill=PatternFill('solid',fgColor='F8CBAD'); warn_fill=PatternFill('solid',fgColor='FFE699')
ok_fill=PatternFill('solid',fgColor='E2EFDA'); inp_fill=PatternFill('solid',fgColor='DDEBF7')
thin=Side(style='thin',color='BFBFBF'); border=Border(left=thin,right=thin,top=thin,bottom=thin)
center=Alignment(horizontal='center',vertical='center'); left=Alignment(horizontal='left',vertical='center',wrap_text=True)

wb=Workbook(); ws=wb.active; ws.title="Menu correcto vs BD"; ws.sheet_view.showGridLines=False
ws.cell(1,1,"Sirope - Menu correcto (carta) frente a lo que hay en el TPV").font=Font(name=FONT,bold=True,size=14,color='1F4E78')
ws.cell(2,1,"Precio bruto (IVA incl.). Azul = precio carta a rellenar. Amarillo = accion (crear/actualizar/duplicado). Rojo = roto o a eliminar. Cabecera roja = bloque a limpiar.").font=Font(name=FONT,italic=True,size=9,color='808080')
headers=["Producto","Precio BD EUR","Precio carta EUR","Dif EUR","Coincide?","Estado / Notas"]
r=4
for c,h in enumerate(headers,1):
    cell=ws.cell(r,c,h); cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=center; cell.border=border
r+=1
for cat in ORDER:
    rows=data.get(cat,[])
    if not rows: continue
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
    cell=ws.cell(r,1,cat); cell.font=catfont; cell.alignment=Alignment(horizontal='left',vertical='center')
    cell.fill=delf if ("eliminar" in cat.lower() or "limpiar" in cat.lower()) else catf
    for c in range(1,7): ws.cell(r,c).border=border
    r+=1
    for (n,bd,ct,note,kind) in rows:
        ws.cell(r,1,n).alignment=left
        cp=ws.cell(r,2,bd); cp.number_format='0.00'; cp.alignment=center
        ca=ws.cell(r,3,ct); ca.number_format='0.00'; ca.alignment=center
        if ct is None: ca.fill=inp_fill
        d=ws.cell(r,4,f'=IF(C{r}="","",IF(B{r}="","",C{r}-B{r}))'); d.number_format='0.00'; d.alignment=center
        ok=ws.cell(r,5,f'=IF(C{r}="","",IF(B{r}="","CREAR",IF(ABS(C{r}-B{r})<0.005,"OK","REVISAR")))'); ok.alignment=center
        ws.cell(r,6,note).alignment=left
        for c in range(1,7): ws.cell(r,c).font=Font(name=FONT,size=10); ws.cell(r,c).border=border
        if kind=='bad':
            for c in range(1,7): ws.cell(r,c).fill=bad_fill
        elif kind=='warn':
            for c in (1,6): ws.cell(r,c).fill=warn_fill
        elif kind=='ok':
            ws.cell(r,6).fill=ok_fill
        r+=1
for c,w in {1:50,2:14,3:15,4:9,5:11,6:44}.items(): ws.column_dimensions[chr(64+c)].width=w
ws.freeze_panes="A5"

# ---- Hoja 2: carta transcripcion (referencia) ----
ws2=wb.create_sheet("Carta (precios nuevos)"); ws2.sheet_view.showGridLines=False
ws2.cell(1,1,"Sirope - Precios NUEVOS de la carta (transcripcion de las fotos)").font=Font(name=FONT,bold=True,size=14,color='1F4E78')
ws2.cell(2,1,"Amarillo = lectura dudosa del manuscrito: VERIFICAR. (Desayunos, Suplementos y Cafe ya confirmados por el usuario.)").font=Font(name=FONT,italic=True,size=9,color='C00000')
hd=["Seccion","Producto (carta)","Precio nuevo EUR","Dudoso?"]
for c,h in enumerate(hd,1):
    cell=ws2.cell(4,c,h); cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=center; cell.border=border
rr=5
confirmed_secs={"CAFE","INFUSIONES","SUPLEMENTOS","DESAYUNOS","BOLLERIA","TOSTADA / MOLLETE","CROISSANT / SANDWICH","BEBIDAS","LICORES"}
for sec,prod,price,dud in carta:
    if sec in confirmed_secs: dud=False
    ws2.cell(rr,1,sec).alignment=center; ws2.cell(rr,2,prod).alignment=left
    pc=ws2.cell(rr,3,price); pc.number_format='0.00'; pc.alignment=center
    ws2.cell(rr,4,"verificar" if dud else "").alignment=center
    for c in range(1,5): ws2.cell(rr,c).font=Font(name=FONT,size=10); ws2.cell(rr,c).border=border
    if dud: ws2.cell(rr,3).fill=warn_fill; ws2.cell(rr,4).fill=warn_fill
    rr+=1
for c,w in {1:22,2:52,3:16,4:12}.items(): ws2.column_dimensions[chr(64+c)].width=w
ws2.freeze_panes="A5"

wb.save("/sessions/nifty-nice-pascal/mnt/outputs/Sirope_precios_TPV_vs_carta.xlsx")
for cat in ORDER: print(f"{cat}: {len(data.get(cat,[]))}")
